"""Create an explicit retrieval-focused copy of the ASK Vera test workbook.

The end-to-end QA workbook contains several kinds of tests: retrieval, guardrail,
conversation, unsupported-source, and rows that need human review. This script
adds explicit columns so retrieval scoring does not rely on fragile text guesses.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]

TEST_TYPE_HEADER = "Test Type"
SOURCE_STATUS_HEADER = "Retrieval Expected Source Status"
CLEANUP_NOTES_HEADER = "Retrieval Cleanup Notes"

OUTPUT_HEADERS = [TEST_TYPE_HEADER, SOURCE_STATUS_HEADER, CLEANUP_NOTES_HEADER]


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalized(value: Any) -> str:
    return re.sub(r"\s+", " ", _text(value).lower())


def _find_header(headers: dict[str, int], candidates: list[str]) -> int | None:
    normalized = {_normalized(name): column for name, column in headers.items()}
    for candidate in candidates:
        if candidate in normalized:
            return normalized[candidate]
    for name, column in headers.items():
        header = _normalized(name)
        if any(candidate in header for candidate in candidates):
            return column
    return None


def _classify_row(category: str, question: str, expected_source: str) -> tuple[str, str, str]:
    category_l = category.lower()
    question_l = question.lower()
    expected_l = expected_source.lower()

    if not expected_source:
        return "needs_review", "missing_expected_source", "No expected source is listed."

    if any(marker in expected_l for marker in ["n/a", "not in kb", "out of scope", "refusal", "consent"]):
        return "guardrail", "not_retrieval", "Expected behavior is block/refusal/consent, not document retrieval."

    if any(marker in category_l for marker in ["prompt injection", "pii", "safety", "guardrail", "medical", "income", "legal", "tax"]):
        return "guardrail", "not_retrieval", "Safety/governance test; score in chatbot QA, not retrieval QA."

    if any(marker in question_l for marker in ["hello", "thank", "tell me more", "turn 1", "turn 2"]):
        return "conversation", "not_retrieval", "Conversation/follow-up behavior test; not a pure retrieval test."

    if any(marker in expected_l for marker in ["customer care", "contact customer care", "support team", "phone number"]):
        return "not_in_document", "source_document_needed", "Needs a confirmed indexed support/contact source."

    if re.search(r"\b(?:sec|section)\.?\s*\d", expected_l):
        return "retrieval", "confirmed_section", "Expected section is provided."

    if ".pdf" in expected_l or "policy" in expected_l:
        return "needs_review", "needs_page_or_section", "Expected document is listed, but page/section should be confirmed."

    return "needs_review", "needs_expected_source_review", "Expected source needs human confirmation."


def _append_or_get_header(ws, header: str) -> int:
    headers = {_text(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
    if header in headers:
        return headers[header]
    column = ws.max_column + 1
    ws.cell(1, column).value = header
    return column


def clean_workbook(input_path: Path, output_path: Path, sheet_name: str) -> dict[str, int]:
    wb = load_workbook(input_path)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet not found: {sheet_name}. Available sheets: {', '.join(wb.sheetnames)}")

    ws = wb[sheet_name]
    headers = {_text(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}

    category_col = _find_header(headers, ["category"])
    question_col = _find_header(headers, ["test question", "question", "prompt", "user question"])
    expected_col = _find_header(
        headers,
        [
            "expected citation / source (ca policy may 2026)",
            "expected citation",
            "expected source",
            "source",
            "document",
            "expected document",
        ],
    )

    if question_col is None or expected_col is None:
        raise RuntimeError("Could not find required question and expected-source columns.")

    output_columns = {header: _append_or_get_header(ws, header) for header in OUTPUT_HEADERS}

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for header, column in output_columns.items():
        cell = ws.cell(1, column)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(column)].width = max(24, len(header) + 4)

    counts: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        question = _text(ws.cell(row, question_col).value)
        if not question:
            continue
        category = _text(ws.cell(row, category_col).value) if category_col else ""
        expected_source = _text(ws.cell(row, expected_col).value)
        test_type, source_status, note = _classify_row(category, question, expected_source)

        ws.cell(row, output_columns[TEST_TYPE_HEADER]).value = test_type
        ws.cell(row, output_columns[SOURCE_STATUS_HEADER]).value = source_status
        ws.cell(row, output_columns[CLEANUP_NOTES_HEADER]).value = note
        counts[test_type] = counts.get(test_type, 0) + 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return counts


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Add explicit retrieval QA labels to the ASK Vera test workbook.")
    parser.add_argument("--xlsx", type=Path, required=True, help="Source ASK Vera workbook.")
    parser.add_argument("--sheet", default="Test Cases", help="Worksheet with test cases.")
    parser.add_argument(
        "--output",
        type=Path,
        default=PROJECT_ROOT / "outputs" / "retrieval_test_review" / "ASK_Vera_Canada_Test_Scenarios_retrieval_labeled.xlsx",
        help="Cleaned workbook output path.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    counts = clean_workbook(args.xlsx, args.output, args.sheet)
    print("Retrieval test plan cleanup complete")
    print("------------------------------------")
    print(f"Input:  {args.xlsx}")
    print(f"Output: {args.output}")
    for test_type, count in sorted(counts.items()):
        print(f"{test_type}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
